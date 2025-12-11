"""
바이낸스 선물 상장 숏 전략 백테스트
=====================================
유니버스: 2023년 이후 바이낸스 선물 상장 코인
진입: 실제 첫 거래 시점 기준 N시간 후 숏
손절/익절/타임아웃으로 청산

사용법: python binance_short_backtest.py
결과: 스크립트와 동일 폴더에 backtest_results.xlsx 생성
"""

import ccxt
import pandas as pd
from datetime import datetime
from typing import Optional, Dict, List
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# 설정값 (필요시 수정)
# ============================================================
ENTRY_DELAY_HOURS = 6       # 상장 후 진입까지 대기 시간
STOP_LOSS_PCT = 0.10        # 손절 10%
TAKE_PROFIT_PCT = 0.40      # 익절 40%
TIMEOUT_HOURS = 72          # 타임아웃 72시간
START_DATE = datetime(2023, 1, 1)  # 이 날짜 이후 상장 코인만
API_DELAY = 1.0             # API 호출 간격 (초)

# 출력 경로 (스크립트와 동일 폴더)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "backtest_results.xlsx")


# ============================================================
# 데이터 수집
# ============================================================
def get_exchange():
    """바이낸스 선물 거래소 객체 생성"""
    return ccxt.binanceusdm({
        'enableRateLimit': True,
        'options': {'defaultType': 'future'}
    })


def get_listings(exchange) -> List[Dict]:
    """2023년 이후 상장된 선물 코인 목록 조회"""
    print("\n📊 바이낸스 선물 상장 정보 수집 중...")
    exchange.load_markets()
    
    listings = []
    for symbol, market in exchange.markets.items():
        if market['quote'] != 'USDT' or not market['active']:
            continue
        
        onboard_date = market.get('info', {}).get('onboardDate')
        if not onboard_date:
            continue
        
        try:
            listing_ts = int(onboard_date)
            if datetime.fromtimestamp(listing_ts / 1000) >= START_DATE:
                listings.append({
                    'symbol': symbol,
                    'base': market['base'],
                    'listing_timestamp': listing_ts
                })
        except (ValueError, TypeError):
            continue
    
    listings.sort(key=lambda x: x['listing_timestamp'])
    print(f"✅ {START_DATE.strftime('%Y-%m-%d')} 이후 상장: {len(listings)}개")
    return listings


def get_ohlcv(exchange, symbol: str, since_ts: int, until_ts: int) -> Optional[pd.DataFrame]:
    """1시간봉 OHLCV 데이터 조회"""
    try:
        ohlcv = exchange.fetch_ohlcv(symbol, timeframe='1h', since=since_ts, limit=1000)
        if not ohlcv:
            return None
        
        df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
        df = df[df['timestamp'] <= until_ts]
        return df if len(df) > 0 else None
    except Exception:
        return None


# ============================================================
# 전략 시뮬레이션
# ============================================================
def simulate_short(df: pd.DataFrame, entry_idx: int) -> Optional[Dict]:
    """
    숏 포지션 시뮬레이션
    - 진입: entry_idx 캔들의 종가
    - 손절: 진입가 * (1 + STOP_LOSS_PCT)
    - 익절: 진입가 * (1 - TAKE_PROFIT_PCT)
    - 타임아웃: entry_idx + TIMEOUT_HOURS
    """
    if entry_idx >= len(df):
        return None
    
    entry_price = df.iloc[entry_idx]['close']
    entry_time = df.iloc[entry_idx]['datetime']
    
    stop_loss_price = entry_price * (1 + STOP_LOSS_PCT)
    take_profit_price = entry_price * (1 - TAKE_PROFIT_PCT)
    timeout_idx = min(entry_idx + TIMEOUT_HOURS, len(df) - 1)
    
    exit_price, exit_time, exit_reason = None, None, None
    max_drawdown, max_profit = 0.0, 0.0
    
    for i in range(entry_idx + 1, timeout_idx + 1):
        if i >= len(df):
            break
        
        high = df.iloc[i]['high']
        low = df.iloc[i]['low']
        
        # 미실현 손익 추적
        max_drawdown = max(max_drawdown, (high - entry_price) / entry_price * 100)
        max_profit = max(max_profit, (entry_price - low) / entry_price * 100)
        
        # 손절 (가격 상승)
        if high >= stop_loss_price:
            exit_price = stop_loss_price
            exit_time = df.iloc[i]['datetime']
            exit_reason = 'STOP_LOSS'
            break
        
        # 익절 (가격 하락)
        if low <= take_profit_price:
            exit_price = take_profit_price
            exit_time = df.iloc[i]['datetime']
            exit_reason = 'TAKE_PROFIT'
            break
    
    # 타임아웃
    if exit_price is None and timeout_idx < len(df):
        exit_price = df.iloc[timeout_idx]['close']
        exit_time = df.iloc[timeout_idx]['datetime']
        exit_reason = 'TIMEOUT'
    
    if exit_price is None:
        return None
    
    pnl_pct = (entry_price - exit_price) / entry_price * 100
    holding_hours = (exit_time - entry_time).total_seconds() / 3600
    
    return {
        'entry_price': entry_price,
        'entry_time': entry_time,
        'exit_price': exit_price,
        'exit_time': exit_time,
        'exit_reason': exit_reason,
        'pnl_pct': pnl_pct,
        'holding_hours': holding_hours,
        'max_drawdown': max_drawdown,
        'max_profit': max_profit
    }


# ============================================================
# 백테스트 실행
# ============================================================
def run_backtest() -> List[Dict]:
    """백테스트 메인 루프"""
    exchange = get_exchange()
    listings = get_listings(exchange)
    results = []
    
    hours_needed = ENTRY_DELAY_HOURS + TIMEOUT_HOURS + 10
    
    print(f"\n🚀 백테스트 시작...\n")
    
    for i, listing in enumerate(listings):
        symbol = listing['symbol']
        base = listing['base']
        
        print(f"[{i+1:3d}/{len(listings)}] {symbol:15s}", end=" | ")
        
        # 데이터 조회
        start_ts = listing['listing_timestamp']
        end_ts = start_ts + hours_needed * 3600 * 1000
        df = get_ohlcv(exchange, symbol, start_ts, end_ts)
        
        if df is None or len(df) < ENTRY_DELAY_HOURS + 10:
            print("❌ 데이터 부족")
            results.append({
                'symbol': symbol, 'base': base,
                'listing_date': None, 'entry_time': None, 'entry_price': None,
                'exit_time': None, 'exit_price': None, 'exit_reason': 'NO_DATA',
                'pnl_pct': None, 'holding_hours': None,
                'max_drawdown': None, 'max_profit': None, 'status': 'SKIPPED'
            })
            time.sleep(API_DELAY)
            continue
        
        # 실제 첫 거래 시점 = 첫 캔들
        actual_listing = df.iloc[0]['datetime']
        
        # 첫 캔들 기준 N시간 후 진입
        trade = simulate_short(df, ENTRY_DELAY_HOURS)
        
        if trade:
            emoji = "✅" if trade['pnl_pct'] > 0 else "🔴"
            print(f"{actual_listing.strftime('%Y-%m-%d %H:%M')} | {emoji} {trade['exit_reason']:11s} | PnL: {trade['pnl_pct']:+7.2f}%")
            
            results.append({
                'symbol': symbol, 'base': base,
                'listing_date': actual_listing,
                'entry_time': trade['entry_time'],
                'entry_price': trade['entry_price'],
                'exit_time': trade['exit_time'],
                'exit_price': trade['exit_price'],
                'exit_reason': trade['exit_reason'],
                'pnl_pct': trade['pnl_pct'],
                'holding_hours': trade['holding_hours'],
                'max_drawdown': trade['max_drawdown'],
                'max_profit': trade['max_profit'],
                'status': 'COMPLETED'
            })
        else:
            print(f"{actual_listing.strftime('%Y-%m-%d %H:%M')} | ⚠️ 미완료")
            results.append({
                'symbol': symbol, 'base': base,
                'listing_date': actual_listing,
                'entry_time': None, 'entry_price': None,
                'exit_time': None, 'exit_price': None,
                'exit_reason': 'INCOMPLETE',
                'pnl_pct': None, 'holding_hours': None,
                'max_drawdown': None, 'max_profit': None, 'status': 'INCOMPLETE'
            })
        
        time.sleep(API_DELAY)
    
    return results


# ============================================================
# 엑셀 리포트 생성
# ============================================================
def create_excel_report(results: List[Dict], filepath: str):
    """엑셀 리포트 생성"""
    wb = Workbook()
    
    # 스타일
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    green_font = Font(color='008000', bold=True)
    red_font = Font(color='FF0000', bold=True)
    bold_font = Font(bold=True, size=11)
    
    def apply_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w
    
    completed = [r for r in results if r['status'] == 'COMPLETED']
    
    # --- Sheet 1: Individual Trades ---
    ws1 = wb.active
    ws1.title = "Individual Trades"
    headers = ['Symbol', 'Base', 'Listing Date', 'Entry Time', 'Entry Price',
               'Exit Time', 'Exit Price', 'Exit Reason', 'PnL (%)',
               'Holding (hrs)', 'Max DD (%)', 'Max Profit (%)', 'Status']
    apply_header(ws1, headers)
    
    for row_idx, r in enumerate(results, 2):
        ws1.cell(row=row_idx, column=1, value=r['symbol'])
        ws1.cell(row=row_idx, column=2, value=r['base'])
        ws1.cell(row=row_idx, column=3, value=r['listing_date'].strftime('%Y-%m-%d %H:%M') if r['listing_date'] else '')
        ws1.cell(row=row_idx, column=4, value=r['entry_time'].strftime('%Y-%m-%d %H:%M') if r['entry_time'] else '')
        ws1.cell(row=row_idx, column=5, value=round(r['entry_price'], 6) if r['entry_price'] else None)
        ws1.cell(row=row_idx, column=6, value=r['exit_time'].strftime('%Y-%m-%d %H:%M') if r['exit_time'] else '')
        ws1.cell(row=row_idx, column=7, value=round(r['exit_price'], 6) if r['exit_price'] else None)
        ws1.cell(row=row_idx, column=8, value=r['exit_reason'])
        
        pnl_cell = ws1.cell(row=row_idx, column=9, value=round(r['pnl_pct'], 2) if r['pnl_pct'] else None)
        if r['pnl_pct'] is not None:
            pnl_cell.font = green_font if r['pnl_pct'] > 0 else red_font
        
        ws1.cell(row=row_idx, column=10, value=round(r['holding_hours'], 1) if r['holding_hours'] else None)
        ws1.cell(row=row_idx, column=11, value=round(r['max_drawdown'], 2) if r['max_drawdown'] else None)
        ws1.cell(row=row_idx, column=12, value=round(r['max_profit'], 2) if r['max_profit'] else None)
        ws1.cell(row=row_idx, column=13, value=r['status'])
    
    set_widths(ws1, [15, 8, 18, 18, 12, 18, 12, 12, 10, 12, 12, 12, 12])
    
    # --- Sheet 2: Exit Reason Summary ---
    ws2 = wb.create_sheet("Exit Reason Summary")
    headers = ['Exit Reason', 'Count', 'Win Rate (%)', 'Avg PnL (%)',
               'Total PnL (%)', 'Avg Holding (hrs)', 'Avg Max DD (%)']
    apply_header(ws2, headers)
    
    row = 2
    for reason in ['STOP_LOSS', 'TAKE_PROFIT', 'TIMEOUT']:
        trades = [r for r in completed if r['exit_reason'] == reason]
        if not trades:
            continue
        pnls = [r['pnl_pct'] for r in trades]
        wins = len([p for p in pnls if p > 0])
        
        ws2.cell(row=row, column=1, value=reason)
        ws2.cell(row=row, column=2, value=len(trades))
        ws2.cell(row=row, column=3, value=round(wins / len(trades) * 100, 1))
        ws2.cell(row=row, column=4, value=round(sum(pnls) / len(pnls), 2))
        ws2.cell(row=row, column=5, value=round(sum(pnls), 2))
        ws2.cell(row=row, column=6, value=round(sum(r['holding_hours'] for r in trades) / len(trades), 1))
        ws2.cell(row=row, column=7, value=round(sum(r['max_drawdown'] for r in trades) / len(trades), 2))
        row += 1
    
    set_widths(ws2, [15, 10, 12, 12, 12, 16, 14])
    
    # --- Sheet 3: Overall Statistics ---
    ws3 = wb.create_sheet("Overall Statistics")
    
    if completed:
        pnls = [r['pnl_pct'] for r in completed]
        wins = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p < 0]
        pf = round(abs(sum(wins) / sum(losses)), 2) if losses and sum(losses) != 0 else 'N/A'
        
        stats = [
            ('트레이드 개요', ''),
            ('총 대상 코인', len(results)),
            ('완료된 트레이드', len(completed)),
            ('스킵된 트레이드', len(results) - len(completed)),
            ('', ''),
            ('승패 분석', ''),
            ('승리 횟수', len(wins)),
            ('패배 횟수', len(losses)),
            ('승률 (%)', round(len(wins) / len(completed) * 100, 2)),
            ('', ''),
            ('수익 분석', ''),
            ('총 수익률 (%)', round(sum(pnls), 2)),
            ('평균 수익률 (%)', round(sum(pnls) / len(pnls), 2)),
            ('최대 단일 수익 (%)', round(max(pnls), 2)),
            ('최대 단일 손실 (%)', round(min(pnls), 2)),
            ('표준편차 (%)', round(pd.Series(pnls).std(), 2)),
            ('', ''),
            ('리스크 지표', ''),
            ('평균 승리 (%)', round(sum(wins) / len(wins), 2) if wins else 0),
            ('평균 손실 (%)', round(sum(losses) / len(losses), 2) if losses else 0),
            ('Profit Factor', pf),
            ('평균 보유 시간 (hrs)', round(sum(r['holding_hours'] for r in completed) / len(completed), 1)),
            ('', ''),
            ('청산 사유별 건수', ''),
            ('STOP_LOSS', len([r for r in completed if r['exit_reason'] == 'STOP_LOSS'])),
            ('TAKE_PROFIT', len([r for r in completed if r['exit_reason'] == 'TAKE_PROFIT'])),
            ('TIMEOUT', len([r for r in completed if r['exit_reason'] == 'TIMEOUT'])),
            ('', ''),
            ('전략 설정', ''),
            ('진입 딜레이 (hrs)', ENTRY_DELAY_HOURS),
            ('손절 (%)', STOP_LOSS_PCT * 100),
            ('익절 (%)', TAKE_PROFIT_PCT * 100),
            ('타임아웃 (hrs)', TIMEOUT_HOURS),
        ]
        
        section_headers = ['트레이드 개요', '승패 분석', '수익 분석', '리스크 지표', '청산 사유별 건수', '전략 설정']
        for row, (label, value) in enumerate(stats, 1):
            cell = ws3.cell(row=row, column=1, value=label)
            if label in section_headers:
                cell.font = bold_font
            ws3.cell(row=row, column=2, value=value)
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 15
    
    # --- Sheet 4: Monthly Performance ---
    ws4 = wb.create_sheet("Monthly Performance")
    headers = ['Month', 'Trades', 'Wins', 'Win Rate (%)',
               'Total PnL (%)', 'Avg PnL (%)', 'Best (%)', 'Worst (%)']
    apply_header(ws4, headers)
    
    monthly: Dict[str, List[float]] = {}
    for r in completed:
        if r['listing_date']:
            month = r['listing_date'].strftime('%Y-%m')
            monthly.setdefault(month, []).append(r['pnl_pct'])
    
    row = 2
    for month in sorted(monthly.keys()):
        pnls = monthly[month]
        wins = len([p for p in pnls if p > 0])
        
        ws4.cell(row=row, column=1, value=month)
        ws4.cell(row=row, column=2, value=len(pnls))
        ws4.cell(row=row, column=3, value=wins)
        ws4.cell(row=row, column=4, value=round(wins / len(pnls) * 100, 1))
        
        total_cell = ws4.cell(row=row, column=5, value=round(sum(pnls), 2))
        total_cell.font = green_font if sum(pnls) > 0 else red_font
        
        ws4.cell(row=row, column=6, value=round(sum(pnls) / len(pnls), 2))
        ws4.cell(row=row, column=7, value=round(max(pnls), 2))
        ws4.cell(row=row, column=8, value=round(min(pnls), 2))
        row += 1
    
    set_widths(ws4, [12, 10, 8, 12, 12, 12, 10, 10])
    
    # --- Sheet 5: Cumulative PnL ---
    ws5 = wb.create_sheet("Cumulative PnL")
    headers = ['Trade #', 'Symbol', 'Entry Date', 'PnL (%)', 'Cumulative PnL (%)']
    apply_header(ws5, headers)
    
    sorted_trades = sorted(completed, key=lambda x: x['entry_time'])
    cumulative = 0.0
    for row_idx, r in enumerate(sorted_trades, 2):
        cumulative += r['pnl_pct']
        ws5.cell(row=row_idx, column=1, value=row_idx - 1)
        ws5.cell(row=row_idx, column=2, value=r['symbol'])
        ws5.cell(row=row_idx, column=3, value=r['entry_time'].strftime('%Y-%m-%d'))
        ws5.cell(row=row_idx, column=4, value=round(r['pnl_pct'], 2))
        ws5.cell(row=row_idx, column=5, value=round(cumulative, 2))
    
    set_widths(ws5, [10, 15, 12, 10, 18])
    
    wb.save(filepath)
    print(f"\n💾 엑셀 저장: {filepath}")


# ============================================================
# 콘솔 요약 출력
# ============================================================
def print_summary(results: List[Dict]):
    """콘솔에 요약 출력"""
    completed = [r for r in results if r['status'] == 'COMPLETED']
    if not completed:
        print("\n⚠️ 완료된 트레이드가 없습니다.")
        return
    
    pnls = [r['pnl_pct'] for r in completed]
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    
    print("\n" + "=" * 60)
    print("📊 백테스트 결과 요약")
    print("=" * 60)
    print(f"총 트레이드: {len(completed)}개")
    print(f"승률: {len(wins)/len(pnls)*100:.1f}% ({len(wins)}승 / {len(losses)}패)")
    print(f"총 수익률: {sum(pnls):.2f}%")
    print(f"평균 수익률: {sum(pnls)/len(pnls):.2f}%")
    print(f"최대 수익: {max(pnls):.2f}%")
    print(f"최대 손실: {min(pnls):.2f}%")
    
    if losses and sum(losses) != 0:
        print(f"Profit Factor: {abs(sum(wins)/sum(losses)):.2f}")
    print("=" * 60)
    
    print("\n📈 청산 사유별:")
    for reason in ['TAKE_PROFIT', 'STOP_LOSS', 'TIMEOUT']:
        trades = [r for r in completed if r['exit_reason'] == reason]
        if trades:
            rpnls = [r['pnl_pct'] for r in trades]
            print(f"  {reason:12s}: {len(trades):3d}건 | 평균 {sum(rpnls)/len(rpnls):+6.2f}%")


# ============================================================
# 메인
# ============================================================
def main():
    print("\n" + "=" * 60)
    print("🚀 바이낸스 선물 상장 숏 전략 백테스트")
    print("=" * 60)
    print(f"📌 전략 설정:")
    print(f"   • 진입: 첫 거래 {ENTRY_DELAY_HOURS}시간 후 숏")
    print(f"   • 손절: {STOP_LOSS_PCT*100:.0f}%")
    print(f"   • 익절: {TAKE_PROFIT_PCT*100:.0f}%")
    print(f"   • 타임아웃: {TIMEOUT_HOURS}시간")
    print(f"   • 유니버스: {START_DATE.strftime('%Y-%m-%d')} 이후 상장")
    print(f"\n📁 결과 저장: {OUTPUT_FILE}")
    print("=" * 60)
    
    results = run_backtest()
    create_excel_report(results, OUTPUT_FILE)
    print_summary(results)
    
    print(f"\n✅ 완료! 결과 파일: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()